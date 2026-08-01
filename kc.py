#!/usr/bin/env python3
"""
Knowledge Chest
===============
A local, free tool that turns a folder of academic-paper PDFs into:

  1. a searchable knowledge base of paper summaries, and
  2. a growing, de-duplicated "near-universe" of datasets used across the papers.

Design goals (yours):
  * PDFs never leave your hard drive. This tool only reads them in place and
    stores the *path* — not the file — in a small database.
  * Everything it extracts lives in ONE portable SQLite file, plus a
    double-click dashboard (knowledge_chest.html) and a spreadsheet
    (knowledge_chest.xlsx) you can drop in Dropbox/Drive to reach anywhere.
  * Free and scales to thousands of papers.

Commands
--------
  python kc.py seed      Load the two sample papers + four sample datasets so you
                         can see the dashboard working before doing any setup.
  python kc.py build     Regenerate the dashboard + spreadsheet from the database.
  python kc.py ingest    Scan PAPERS_DIR, extract every new paper with Claude,
                         store it, then rebuild the dashboard + spreadsheet.
  python kc.py stats     Print counts.

Setup for `ingest` (the only part that needs the internet):
  pip install -r requirements.txt
  set ANTHROPIC_API_KEY   (get one at https://console.anthropic.com )
  Put your PDFs in the ./papers folder (or point PAPERS_DIR at your library).

Everything else — seed, build, browsing the dashboard — works offline with no key.
"""

import os
import re
import io
import sys
import json
import time
import base64
import hashlib
import sqlite3
import zipfile
import datetime as _dt
from pathlib import Path

# --------------------------------------------------------------------------- #
# CONFIG  — override any of these with an environment variable of the same name
# --------------------------------------------------------------------------- #
HERE = Path(__file__).resolve().parent

def _load_dotenv():
    """Load KEY=value lines from a local .env file (no dependency)."""
    p = HERE / ".env"
    if not p.exists():
        return
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

_load_dotenv()

def _env(name, default):
    return os.environ.get(name, default)

# Where YOUR papers live. Point this at any folder on your computer via .env,
# e.g. PAPERS_DIR=/Users/you/Documents/Papers  — it is read locally and never uploaded.
PAPERS_DIR  = Path(_env("PAPERS_DIR",  str(HERE / "papers"))).expanduser()
# After a paper is catalogued it is renamed to its standardized name and MOVED here.
PROCESSED_DIR = Path(_env("PROCESSED_DIR", str(HERE / "processed"))).expanduser()
# Set KC_MOVE=0 to catalogue papers in place instead of moving them.
MOVE_PROCESSED = _env("KC_MOVE", "1") == "1"
# Keep filenames short enough for Windows' path limits.
MAX_FILENAME = int(_env("KC_MAX_FILENAME", "150"))
# Folder (on your own computer) that holds the renamed PDFs. The dashboard adds
# an "Open PDF" link to each paper pointing here. Set "" to hide the links.
PAPERS_LINK_BASE = _env("KC_LINK_BASE",
    "file:///D:/Dropbox/_Bugra's Folder/Research/4. Knowledge Chest/Papers/Processed/")
# The dashboard + spreadsheet go here. GitHub Pages serves this folder ("/docs").
OUTPUT_DIR  = Path(_env("OUTPUT_DIR",  str(HERE / "docs"))).expanduser()
# The database (your source of truth) lives here, kept out of the served folder.
DB_PATH     = Path(_env("DB_PATH",     str(HERE / "data" / "knowledge_chest.db"))).expanduser()

# Model: Haiku is cheap and good for extraction (~1-2 cents/paper). Switch to
# "claude-sonnet-5" if you want higher-quality summaries at a bit more cost.
MODEL       = _env("KC_MODEL", "claude-haiku-4-5-20251001")

# How much paper text to send per call (chars). ~4 chars/token, so 32000 ≈ 8k tokens.
MAX_TEXT_CHARS = int(_env("KC_MAX_TEXT_CHARS", "32000"))
# For scanned papers, how many page images to send (each costs more than text).
MAX_SCAN_PAGES = int(_env("KC_MAX_SCAN_PAGES", "6"))

# Datasets so ubiquitous we do NOT want them cluttering the "near-universe".
COMMON_DATASETS = [
    "sec edgar", "edgar", "compustat", "i/b/e/s", "ibes", "crsp",
    "datastream", "worldscope", "thomson reuters", "bloomberg terminal",
]

# --------------------------------------------------------------------------- #
# DATABASE
# --------------------------------------------------------------------------- #
SCHEMA = """
CREATE TABLE IF NOT EXISTS papers (
    id              INTEGER PRIMARY KEY,
    file_hash       TEXT UNIQUE,
    file_path       TEXT,
    orig_filename   TEXT,
    std_name        TEXT,
    journal         TEXT,
    is_working_paper INTEGER DEFAULT 0,
    year            INTEGER,
    authors_json    TEXT,
    title           TEXT,
    summary         TEXT,
    logical_flow    TEXT,
    research_design TEXT,
    categories_json TEXT,
    shock           TEXT,
    missing_notes   TEXT,
    added_at        TEXT
);
CREATE TABLE IF NOT EXISTS datasets (
    id            INTEGER PRIMARY KEY,
    dedup_key     TEXT UNIQUE,
    provider      TEXT,
    product       TEXT,
    aliases       TEXT,
    description   TEXT,
    access_type   TEXT,
    delivery      TEXT,
    topic_tags    TEXT,
    observations  TEXT,      -- JSON list of description snippets contributed by papers
    first_seen_at TEXT
);
CREATE TABLE IF NOT EXISTS paper_datasets (
    paper_id   INTEGER,
    dataset_id INTEGER,
    PRIMARY KEY (paper_id, dataset_id)
);
"""

def get_conn():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    # Migrate older databases that predate a column.
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(datasets)")}
    if "observations" not in cols:
        conn.execute("ALTER TABLE datasets ADD COLUMN observations TEXT")
        conn.commit()
    return conn

def _norm(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())

def dataset_dedup_key(provider, product):
    return _norm(provider) + " :: " + _norm(product)

def _union_semis(*chunks):
    """Union of ';'/','-separated tokens across all inputs, order-stable-ish."""
    parts = []
    seen = set()
    for chunk in chunks:
        for x in re.split(r"[;,]", chunk or ""):
            x = x.strip()
            if x and x != "—" and x.lower() not in seen:
                seen.add(x.lower()); parts.append(x)
    return "; ".join(parts)

def _dedup_snippets(snippets):
    out, seen = [], set()
    for s in snippets:
        s = (s or "").strip()
        k = _norm(s)
        if s and k not in seen:
            seen.add(k); out.append(s)
    return out

def synthesize_description(name, snippets, client):
    """Fold every paper's take on a dataset into one clear canonical description."""
    joined = "\n".join(f"- {s}" for s in snippets)
    msg = (f"Dataset: {name}\n\n"
           f"Descriptions gathered from papers that use it:\n{joined}\n\n"
           "Write ONE clear, self-contained description of this dataset in 1-3 sentences. "
           "Cover what the data contains, its coverage (geography and time period if mentioned), "
           "the unit of observation, and how it is accessed. Merge every distinct detail without "
           "repeating, and read naturally. Do not invent anything not present above. "
           "Return only the description text.")
    resp = client.messages.create(model=MODEL, max_tokens=300,
                                  messages=[{"role": "user", "content": msg}])
    return "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()

def compile_description(snippets, client=None):
    """Canonical description from all contributed snippets.
    Uses Claude when a client is available (clearer); otherwise a clean deterministic merge."""
    uniq = _dedup_snippets(snippets)
    if not uniq:
        return ""
    if len(uniq) == 1:
        return uniq[0]
    if client is not None:
        try:
            return synthesize_description("dataset", uniq, client)
        except Exception:
            pass
    # Deterministic fallback: longest as the base, append genuinely new sentences.
    uniq.sort(key=len, reverse=True)
    base = uniq[0]
    for s in uniq[1:]:
        if _norm(s) not in _norm(base):
            base = base.rstrip(". ") + ". " + s
    return base

def upsert_dataset(conn, d, client=None, source_name=None):
    """Insert a dataset, or accumulate new information into an existing one; return id.
    As papers pile in, aliases / tags / access / delivery are unioned and the
    description is recompiled from every contributed snippet."""
    provider = (d.get("provider") or "").strip()
    product  = (d.get("product") or "").strip()
    if not provider and not product:
        return None
    # Safety net: drop ubiquitous sources even if the model returns them.
    if _norm(provider) in COMMON_DATASETS or _norm(product) in COMMON_DATASETS:
        return None

    new_snip = (d.get("description") or "").strip()
    key = dataset_dedup_key(provider, product)
    row = conn.execute("SELECT * FROM datasets WHERE dedup_key = ?", (key,)).fetchone()

    if row:
        obs = json.loads(row["observations"] or "[]")
        adds_info = new_snip and _norm(new_snip) not in {_norm(o) for o in obs} \
                    and _norm(new_snip) not in _norm(row["description"] or "")
        if new_snip and adds_info:
            obs.append(new_snip)
        aliases   = _union_semis(row["aliases"],   d.get("aliases"))
        topic     = _union_semis(row["topic_tags"], d.get("topic_tags"))
        access    = _union_semis(row["access_type"], d.get("access_type"))
        delivery  = _union_semis(row["delivery"],   d.get("delivery"))
        # Recompile the description only when a paper actually added something new.
        desc = compile_description(obs, client) if adds_info else (row["description"] or "")
        conn.execute(
            "UPDATE datasets SET aliases=?, description=?, access_type=?, delivery=?, "
            "topic_tags=?, observations=? WHERE id=?",
            (aliases, desc, access, delivery, topic, json.dumps(obs), row["id"]))
        return row["id"]

    obs = [new_snip] if new_snip else []
    cur = conn.execute(
        "INSERT INTO datasets (dedup_key, provider, product, aliases, description, "
        "access_type, delivery, topic_tags, observations, first_seen_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?)",
        (key, provider, product, _union_semis(d.get("aliases")), new_snip,
         _union_semis(d.get("access_type")), _union_semis(d.get("delivery")),
         _union_semis(d.get("topic_tags")), json.dumps(obs), _dt.date.today().isoformat()))
    return cur.lastrowid

def save_paper(conn, rec, file_hash, file_path, orig_filename, client=None):
    cur = conn.execute(
        "INSERT INTO papers (file_hash, file_path, orig_filename, std_name, journal, "
        "is_working_paper, year, authors_json, title, summary, logical_flow, "
        "research_design, categories_json, shock, missing_notes, added_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (file_hash, str(file_path), orig_filename, rec["std_name"], rec.get("journal"),
         1 if rec.get("is_working_paper") else 0, rec.get("year"),
         json.dumps(rec.get("authors") or []), rec.get("title"), rec.get("summary"),
         rec.get("logical_flow"), rec.get("research_design"),
         json.dumps(rec.get("categories") or []),
         (json.dumps(rec["shock"]) if rec.get("shock") else None), rec.get("missing_notes"),
         _dt.datetime.now().isoformat(timespec="seconds")),
    )
    paper_id = cur.lastrowid
    for d in (rec.get("datasets") or []):
        ds_id = upsert_dataset(conn, d, client=client, source_name=rec.get("std_name"))
        if ds_id:
            conn.execute("INSERT OR IGNORE INTO paper_datasets VALUES (?,?)",
                         (paper_id, ds_id))
    conn.commit()
    return paper_id

# --------------------------------------------------------------------------- #
# STANDARDIZED NAME  ("Journal - Year - Author(s) - Title")
# --------------------------------------------------------------------------- #
def format_authors(last_names):
    ln = [a.strip() for a in (last_names or []) if a and a.strip()]
    # If the list already carries an "et al." marker, collapse to first author.
    if any(re.fullmatch(r"et\.?\s*al\.?", a, re.I) for a in ln):
        first = next((a for a in ln if not re.fullmatch(r"et\.?\s*al\.?", a, re.I)), "Unknown")
        return f"{first} et al."
    if not ln:
        return "Unknown"
    if len(ln) == 1:
        return ln[0]
    if len(ln) == 2:
        return f"{ln[0]} and {ln[1]}"
    return f"{ln[0]} et al."

_BAD_FS = re.compile(r'[<>:"/\\|?*]')

def build_std_name(rec):
    journal = "Working Paper" if rec.get("is_working_paper") else (rec.get("journal") or "Unknown Journal")
    year = rec.get("year") or "n.d."
    authors = format_authors(rec.get("authors"))
    title = (rec.get("title") or "Untitled").strip().rstrip(".")
    name = f"{journal} - {year} - {authors} - {title}"
    name = _BAD_FS.sub("", name)          # strip characters illegal in filenames
    return re.sub(r"\s+", " ", name).strip()

# --------------------------------------------------------------------------- #
# READING PDFs  (text first; fall back to page images for scanned papers)
# --------------------------------------------------------------------------- #
def sha256_of(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()

def read_pdf(path):
    """Return ('text', str) for born-digital PDFs, or ('images', [jpeg_bytes,...])
    for scanned ones. Also handles the odd 'zip of page images' case."""
    path = Path(path)
    head = open(path, "rb").read(4)

    # Case 1: a ZIP holding page images (some scanners/exporters do this).
    if head[:2] == b"PK":
        try:
            z = zipfile.ZipFile(path)
            imgs = sorted([n for n in z.namelist()
                           if n.lower().endswith((".jpg", ".jpeg", ".png"))],
                          key=lambda n: _natkey(n))
            if imgs:
                return "images", [z.read(n) for n in imgs[:MAX_SCAN_PAGES]]
        except Exception:
            pass

    # Case 2: a real PDF — try text with PyMuPDF.
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise SystemExit("Please `pip install -r requirements.txt` (needs PyMuPDF).")
    doc = fitz.open(str(path))
    text = "".join(page.get_text() for page in doc)
    if len(text.strip()) >= 400:                      # enough real text -> digital
        return "text", text

    # Case 3: scanned PDF — render the first pages to JPEG for Claude's vision.
    imgs = []
    for page in doc[:MAX_SCAN_PAGES]:
        pix = page.get_pixmap(dpi=150)
        imgs.append(pix.tobytes("jpeg"))
    return "images", imgs

def _natkey(s):
    return [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", s)]

# --------------------------------------------------------------------------- #
# EXTRACTION  (Claude turns a paper into structured fields)
# --------------------------------------------------------------------------- #
SYSTEM_PROMPT = """You are Knowledge Chest, a meticulous academic-paper cataloguer for economics, finance, and accounting research. You convert one paper into a fixed set of fields so that entries stay comparable across a whole library. You never invent details: when the journal, year, research design, a dataset provider, or any other key fact cannot be determined from the text, you say it is missing rather than guessing.

Return ONLY a single JSON object (no prose, no markdown fences) with exactly these keys:

{
  "journal": string|null,            // publication name; null if unknown
  "is_working_paper": boolean,       // true for unpublished manuscripts/working papers
  "year": integer|null,              // publication/working-paper year; null if unknown
  "authors": [string],               // author LAST names only, in order
  "title": string,                   // full paper title
  "summary": string,                 // 4-5 sentences: research question, main argument,
                                     //   approach, principal findings, contribution.
                                     //   Do NOT copy the abstract; do NOT add outside interpretation.
  "logical_flow": string,            // A DETAILED paragraph (6-9 sentences) tracing how the
                                     //   argument is built: the motivating problem or puzzle, the
                                     //   key conceptual/theoretical mechanisms the authors propose,
                                     //   the assumptions that connect them, the predictions or
                                     //   hypotheses that follow, and how each step sets up the next.
                                     //   Explain the economic intuition, not just the sequence.
                                     //   Emphasize theory/argument development over validation;
                                     //   do not dwell on robustness/placebo tests.
  "research_design": string,         // A DETAILED explanation (4-6 sentences) of the PRIMARY design
                                     //   (e.g., difference-in-differences, RD, IV, RCT, event study,
                                     //   structural estimation, descriptive, theory model). Name it,
                                     //   then explain the source of identifying variation (or the
                                     //   model's key ingredients), the unit and level of analysis,
                                     //   the main comparison/estimating equation IN WORDS, and any
                                     //   secondary method used to support it. Not a dataset inventory.
  "categories": [string],            // up to THREE standardized fields, e.g.
                                     //   ["Labor Economics","Regulation","Political Economy"].
                                     //   Avoid many broad/loosely-related labels.
  "datasets": [                      // only DISTINCTIVE / non-standard data sources
    {
      "provider": string,            // provider or dataset name (e.g., "IRS", "Orbital Insight")
      "product": string|null,        // specific product/table if named, else null
      "description": string,         // a clear, self-contained description of the dataset:
                                     //   WHAT it contains, its COVERAGE (geography and time period
                                     //   if stated), the UNIT OF OBSERVATION (e.g., firm-year,
                                     //   county-day, transaction), and HOW it is accessed.
                                     //   1-3 sentences. Only facts stated in this paper.
      "access_type": string|null,    // "Public" | "Proprietary" | "Restricted" if stated, else null
      "delivery": string|null,       // "API" | "Bulk" | "Web" | etc. if stated, else null
      "topic_tags": string|null      // 2-4 short semicolon-separated tags if useful, else null
    }
  ],
  "no_nonstandard_datasets": boolean,// true if the paper relies only on common sources
  "shock": {                         // the main quasi-exogenous/exogenous shock or natural
                                     //   experiment the paper exploits for identification, or null
                                     //   if the paper uses none (e.g., pure theory, descriptive,
                                     //   structural-only, or survey work).
     "name": string,                 //   short label, e.g. "Alice v. CLS Bank Supreme Court decision"
     "type": string,                 //   e.g. "Supreme Court decision", "policy reform", "natural
                                     //     experiment", "monetary policy shock", "instrument/shift-share"
     "what": string                  //   1-2 sentences: what the shock is and why it is plausibly
                                     //     exogenous to the outcome being studied.
  } ,                                //   Use null (not an object) when there is no such shock.
  "missing_notes": string|null       // note anything you could not determine reliably, else null
}

EXCLUDE these ubiquitous sources from "datasets" (they are assumed, not distinctive):
SEC EDGAR, Compustat, I/B/E/S, CRSP, Datastream, Worldscope, and Bloomberg terminal pulls.
If after excluding them no distinctive dataset remains, set "datasets": [] and
"no_nonstandard_datasets": true."""

def _client():
    try:
        import anthropic
    except ImportError:
        raise SystemExit("Please `pip install -r requirements.txt` (needs the anthropic SDK).")
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise SystemExit("Set ANTHROPIC_API_KEY before running `ingest`. "
                         "Get a key at https://console.anthropic.com")
    return anthropic.Anthropic(api_key=key)

def _parse_json(text):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\n?|\n?```$", "", text).strip()
    # Grab the outermost {...} in case the model added stray characters.
    m = re.search(r"\{.*\}", text, re.DOTALL)
    return json.loads(m.group(0) if m else text)

def extract_fields(kind, payload, client=None):
    """kind = 'text' | 'images'. Returns the parsed record dict (+ std_name)."""
    client = client or _client()
    if kind == "text":
        content = [{"type": "text",
                    "text": "Catalogue this paper:\n\n" + payload[:MAX_TEXT_CHARS]}]
    else:
        content = [{"type": "text",
                    "text": "This paper is scanned. Read the page images and catalogue it:"}]
        for img in payload:
            content.append({"type": "image", "source": {
                "type": "base64", "media_type": "image/jpeg",
                "data": base64.b64encode(img).decode()}})

    resp = client.messages.create(
        model=MODEL, max_tokens=1500, system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    rec = _parse_json(text)
    rec["std_name"] = build_std_name(rec)
    return rec

# --------------------------------------------------------------------------- #
# INGEST
# --------------------------------------------------------------------------- #
def ingest():
    if not PAPERS_DIR.exists():
        PAPERS_DIR.mkdir(parents=True, exist_ok=True)
        print(f"Created {PAPERS_DIR}. Put your PDFs there and run `ingest` again.")
        return
    processed_resolved = PROCESSED_DIR.resolve()
    pdfs = sorted(p for p in PAPERS_DIR.rglob("*")
                  if p.suffix.lower() == ".pdf" and p.is_file()
                  and processed_resolved not in p.resolve().parents)  # never re-read Processed
    if not pdfs:
        print(f"No PDFs found in {PAPERS_DIR}.")
        return

    conn = get_conn()
    client = _client()                       # fail fast if key/SDK missing
    done = {r["file_hash"] for r in conn.execute("SELECT file_hash FROM papers")}
    added = skipped = failed = 0

    for i, pdf in enumerate(pdfs, 1):
        h = sha256_of(pdf)
        if h in done:
            skipped += 1
            continue
        print(f"[{i}/{len(pdfs)}] {pdf.name} ...", end=" ", flush=True)
        try:
            kind, payload = read_pdf(pdf)
            rec = extract_fields(kind, payload, client)
            save_paper(conn, rec, h, pdf, pdf.name, client=client)
            done.add(h)
            added += 1
            # Rename to the standardized name and move OUT of the inbox — only now
            # that it is safely catalogued. Failed papers stay put for a retry.
            final = move_to_processed(conn, pdf, rec["std_name"])
            where = f" -> Processed/{final.name}" if final != pdf else ""
            print("OK ->", rec["std_name"][:70] + where)
        except Exception as e:                # never let one bad file stop the run
            failed += 1
            print("FAILED (left in inbox):", str(e)[:100])
        time.sleep(0.2)                        # be gentle on rate limits

    conn.close()
    print(f"\nAdded {added}, skipped {skipped} (already done), failed {failed}.")
    if failed:
        print(f"Failed papers remain in {PAPERS_DIR} for you to check or retry.")
    build()

def move_to_processed(conn, pdf, std_name):
    """Rename to the standardized name and move into PROCESSED_DIR.
    Returns the final Path (unchanged if moving is disabled). Collisions get a
    numeric suffix so nothing is overwritten."""
    if not MOVE_PROCESSED:
        return pdf
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    stem = std_name[:MAX_FILENAME].rstrip(" .") or "Untitled"
    target = PROCESSED_DIR / (stem + ".pdf")
    n = 2
    while target.exists() and target.resolve() != Path(pdf).resolve():
        target = PROCESSED_DIR / f"{stem} ({n}).pdf"
        n += 1
    try:
        import shutil
        shutil.move(str(pdf), str(target))
    except Exception as e:
        print(f"(kept in inbox — move failed: {str(e)[:60]})", end=" ")
        return pdf
    conn.execute("UPDATE papers SET file_path=? WHERE file_path=?",
                 (str(target), str(pdf)))
    conn.commit()
    return target

# --------------------------------------------------------------------------- #
# BUILD OUTPUTS  (dashboard + spreadsheet)
# --------------------------------------------------------------------------- #
def _fetch_all(conn):
    papers = []
    for r in conn.execute("SELECT * FROM papers ORDER BY year DESC, id DESC"):
        ds = conn.execute(
            "SELECT d.provider, d.product FROM datasets d "
            "JOIN paper_datasets pd ON pd.dataset_id=d.id WHERE pd.paper_id=?",
            (r["id"],)).fetchall()
        papers.append({
            "std_name": r["std_name"], "journal": r["journal"],
            "year": r["year"], "title": r["title"],
            "authors": json.loads(r["authors_json"] or "[]"),
            "summary": r["summary"], "logical_flow": r["logical_flow"],
            "research_design": r["research_design"],
            "categories": json.loads(r["categories_json"] or "[]"),
            "shock": (json.loads(r["shock"]) if r["shock"] else None),
            "missing_notes": r["missing_notes"], "file_path": r["file_path"],
            "datasets": [f"{x['provider']}" + (f" — {x['product']}" if x["product"] else "")
                         for x in ds],
        })
    datasets = []
    for d in conn.execute("SELECT * FROM datasets ORDER BY provider COLLATE NOCASE"):
        users = conn.execute(
            "SELECT p.std_name FROM papers p JOIN paper_datasets pd ON pd.paper_id=p.id "
            "WHERE pd.dataset_id=? ORDER BY p.year DESC", (d["id"],)).fetchall()
        datasets.append({
            "provider": d["provider"], "product": d["product"],
            "aliases": d["aliases"], "description": d["description"],
            "access_type": d["access_type"], "delivery": d["delivery"],
            "topic_tags": d["topic_tags"],
            "sources": len(json.loads(d["observations"] or "[]")),
            "used_by": [u["std_name"] for u in users],
        })
    return papers, datasets

def build():
    conn = get_conn()
    papers, datasets = _fetch_all(conn)
    conn.close()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    _write_html(papers, datasets)
    _write_xlsx(papers, datasets)
    print(f"Dashboard : {OUTPUT_DIR / 'index.html'}")
    print(f"Spreadsheet: {OUTPUT_DIR / 'knowledge_chest.xlsx'}")
    print(f"({len(papers)} papers, {len(datasets)} datasets)")

def _write_xlsx(papers, datasets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5D50")
    body_font = Font(name="Arial")

    ws = wb.active
    ws.title = "Papers"
    pcols = ["Standardized Name", "Journal", "Year", "Authors", "Title",
             "Summary", "Logical Flow", "Research Design", "Categories",
             "Datasets Used", "Missing Notes", "File Path"]
    ws.append(pcols)
    for p in papers:
        ws.append([p["std_name"], p["journal"], p["year"], ", ".join(p["authors"]),
                   p["title"], p["summary"], p["logical_flow"], p["research_design"],
                   ", ".join(p["categories"]), "; ".join(p["datasets"]),
                   p["missing_notes"], p["file_path"]])

    ws2 = wb.create_sheet("Dataset Universe")
    dcols = ["Provider", "Product", "Aliases/Historical Names", "Description",
             "Access Type", "Delivery", "Topic Tags", "# Papers", "Used By (papers)"]
    ws2.append(dcols)
    for d in datasets:
        ws2.append([d["provider"], d["product"], d["aliases"], d["description"],
                    d["access_type"], d["delivery"], d["topic_tags"],
                    len(d["used_by"]), "; ".join(d["used_by"])])

    ws3 = wb.create_sheet("Exogenous Shocks")
    scols = ["Paper", "Authors", "Shock / Natural Experiment", "Type", "What the Shock Is"]
    ws3.append(scols)
    for p in papers:
        s = p.get("shock")
        if not s:
            continue
        ws3.append([p["title"], ", ".join(p["authors"]),
                    s.get("name"), s.get("type"), s.get("what")])

    widths = {"Papers": [34,20,7,18,34,52,52,40,26,26,26,40],
              "Dataset Universe": [22,26,24,50,12,12,26,9,44],
              "Exogenous Shocks": [40,20,32,20,64]}
    for sheet, ws_ in (("Papers", ws), ("Dataset Universe", ws2), ("Exogenous Shocks", ws3)):
        for j, cell in enumerate(ws_[1], 0):
            cell.font = head_font; cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for col_i, w in enumerate(widths[sheet], 1):
            ws_.column_dimensions[chr(64+col_i) if col_i <= 26 else "A"].width = w
        for row in ws_.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_.freeze_panes = "A2"
        ws_.auto_filter.ref = ws_.dimensions
    wb.save(OUTPUT_DIR / "knowledge_chest.xlsx")

# ---- dashboard (single self-contained HTML file, double-click to open) ----- #
def _write_html(papers, datasets):
    data = json.dumps({"papers": papers, "datasets": datasets}, ensure_ascii=False)
    html = _HTML_TEMPLATE.replace("/*__DATA__*/null", data)
    html = html.replace('/*__LINKBASE__*/""', json.dumps(PAPERS_LINK_BASE))
    (OUTPUT_DIR / "index.html").write_text(html, encoding="utf-8")

_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Knowledge Chest</title>
<style>
  :root{
    --paper:#FBFAF7; --ink:#1b1a17; --muted:#6f6a60; --line:#e4ded2;
    --card:#ffffff; --accent:#2f5d50; --accent-soft:#e7efe9; --amber:#9a6a1c;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--paper);color:var(--ink);
    font-family:Georgia,"Iowan Old Style",serif;line-height:1.5}
  .wrap{max-width:1060px;margin:0 auto;padding:28px 22px 80px}
  header{border-bottom:2px solid var(--ink);padding-bottom:14px;margin-bottom:6px}
  h1{font-size:30px;margin:0;letter-spacing:.3px}
  .sub{font:12px/1.4 ui-monospace,"SFMono-Regular",Menlo,monospace;
    color:var(--muted);text-transform:uppercase;letter-spacing:.14em;margin-top:6px}
  .tabs{display:flex;gap:6px;margin:18px 0 14px}
  .tab{font:12px/1 ui-monospace,Menlo,monospace;letter-spacing:.12em;text-transform:uppercase;
    padding:9px 14px;border:1px solid var(--line);background:var(--card);cursor:pointer;
    color:var(--muted);border-radius:2px}
  .tab.on{background:var(--ink);color:var(--paper);border-color:var(--ink)}
  .controls{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:16px}
  input[type=search],select{font:15px Georgia,serif;padding:9px 11px;border:1px solid var(--line);
    background:var(--card);border-radius:2px;color:var(--ink)}
  input[type=search]{flex:1;min-width:220px}
  input[type=search]:focus,select:focus{outline:2px solid var(--accent);outline-offset:1px}
  .count{font:12px ui-monospace,Menlo,monospace;color:var(--muted);letter-spacing:.06em}
  .card{background:var(--card);border:1px solid var(--line);border-left:3px solid var(--accent);
    border-radius:3px;padding:16px 18px;margin-bottom:14px;box-shadow:0 1px 0 rgba(0,0,0,.02)}
  .callno{font:11px/1.5 ui-monospace,Menlo,monospace;color:var(--accent);letter-spacing:.04em;
    text-transform:uppercase;margin-bottom:6px;word-break:break-word}
  .ptitle{font-size:19px;font-weight:700;margin:0 0 4px;cursor:pointer}
  .ptitle:hover{color:var(--accent)}
  .meta{font:12px ui-monospace,Menlo,monospace;color:var(--muted);margin-bottom:8px}
  .cats{margin:8px 0 2px}
  .chip{display:inline-block;font:11px ui-monospace,Menlo,monospace;background:var(--accent-soft);
    color:var(--accent);padding:2px 8px;border-radius:20px;margin:0 6px 6px 0;letter-spacing:.03em}
  .chip.data{background:#f3ecdd;color:var(--amber)}
  .fchip{cursor:pointer;display:inline-block;font:11px ui-monospace,Menlo,monospace;
    background:var(--paper);color:var(--accent);border:1px solid var(--line);
    padding:3px 10px;border-radius:20px;margin:0 6px 6px 0;letter-spacing:.03em;user-select:none}
  .fchip:hover{border-color:var(--accent)}
  .fchip.on{background:var(--accent);color:var(--paper);border-color:var(--accent)}
  .fchip .n{opacity:.55;margin-left:6px}
  .themebar,.tagbar{margin:0 0 6px}
  .barlabel{display:inline-block;font:10px ui-monospace,Menlo,monospace;text-transform:uppercase;
    letter-spacing:.12em;color:var(--muted);margin:3px 8px 4px 0;vertical-align:middle}
  .field{margin:10px 0}
  .flabel{font:11px ui-monospace,Menlo,monospace;text-transform:uppercase;letter-spacing:.12em;
    color:var(--muted);margin-bottom:2px}
  details>summary{cursor:pointer;font:12px ui-monospace,Menlo,monospace;color:var(--accent);
    letter-spacing:.06em;list-style:none;margin-top:4px}
  details>summary::-webkit-details-marker{display:none}
  details[open]>summary{margin-bottom:8px}
  .path{font:11px ui-monospace,Menlo,monospace;color:var(--muted);word-break:break-all;margin-top:8px}
  .path a{color:var(--accent)}
  .note{font-style:italic;color:var(--amber);font-size:13px}
  .empty{color:var(--muted);text-align:center;padding:40px;font-style:italic}
  .ds .used{font:12px Georgia,serif;color:var(--muted);margin-top:6px}
  footer{margin-top:34px;border-top:1px solid var(--line);padding-top:12px;
    font:11px ui-monospace,Menlo,monospace;color:var(--muted);letter-spacing:.05em}
  @media (max-width:560px){h1{font-size:24px}}
</style></head>
<body><div class="wrap">
  <header>
    <h1>Knowledge Chest</h1>
    <div class="sub">A catalogue of what you've read &nbsp;·&nbsp; economics · finance · accounting</div>
  </header>
  <div class="tabs">
    <div class="tab on" data-tab="papers">Papers</div>
    <div class="tab" data-tab="datasets">Dataset Universe</div>
  </div>

  <section id="papers-view">
    <div class="controls">
      <input id="q" type="search" placeholder="Search title, summary, author, design, dataset…" aria-label="Search papers">
      <select id="fcat"><option value="">All categories</option></select>
      <select id="fjour"><option value="">All journals</option></select>
      <select id="fyear"><option value="">All years</option></select>
      <span class="count" id="pcount"></span>
    </div>
    <div id="papers"></div>
  </section>

  <section id="datasets-view" style="display:none">
    <div class="controls">
      <input id="dq" type="search" placeholder="Search provider, description, topic…" aria-label="Search datasets">
      <span class="count" id="dcount"></span>
    </div>
    <div id="dthemes" class="themebar"></div>
    <div id="dtags" class="tagbar"></div>
    <div id="datasets" class="ds"></div>
  </section>

  <footer>Searchable catalogue · click a title to expand its logical flow, research design, and datasets</footer>
</div>
<script>
const DB = /*__DATA__*/null;
const LINK_BASE = /*__LINKBASE__*/"";
const esc = s => (s==null?"":String(s)).replace(/[&<>"]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));
const papers = DB.papers, datasets = DB.datasets;

// filter option lists
const cats = [...new Set(papers.flatMap(p=>p.categories))].sort();
const jours = [...new Set(papers.map(p=>p.journal).filter(Boolean))].sort();
const years = [...new Set(papers.map(p=>p.year).filter(Boolean))].sort((a,b)=>b-a);
const fill=(id,arr)=>{const s=document.getElementById(id);arr.forEach(v=>{const o=document.createElement("option");o.value=v;o.textContent=v;s.appendChild(o);});};
fill("fcat",cats); fill("fjour",jours); fill("fyear",years);

function paperCard(p){
  const auth = p.authors&&p.authors.length?esc(p.authors.join(", ")):"Unknown";
  const cats = p.categories.map(c=>`<span class="chip">${esc(c)}</span>`).join("");
  const data = p.datasets.map(d=>`<span class="chip data">${esc(d)}</span>`).join("") || `<span class="note">no non-standard datasets</span>`;
  const note = p.missing_notes? `<div class="field"><span class="note">Note: ${esc(p.missing_notes)}</span></div>`:"";
  let fileLink = "";
  if (LINK_BASE) {
    const fn = p.std_name.slice(0,150).replace(/[ .]+$/,"") + ".pdf";
    fileLink = `<div class="path"><a href="${encodeURI(LINK_BASE + fn)}">Open PDF ↗</a></div>`;
  }
  return `<div class="card">
    <div class="ptitle" onclick="this.parentElement.querySelector('.detail').toggleAttribute('hidden')">${esc(p.title||"Untitled")}</div>
    <div class="meta">${auth} &nbsp;·&nbsp; ${esc(p.journal||"Working paper")} &nbsp;·&nbsp; ${esc(p.year||"n.d.")}</div>
    <div class="cats">${cats}</div>
    <div class="field"><div class="flabel">Summary</div>${esc(p.summary||"")}</div>
    <div class="detail" hidden>
      <div class="field"><div class="flabel">Logical flow</div>${esc(p.logical_flow||"")}</div>
      <div class="field"><div class="flabel">Research design</div>${esc(p.research_design||"")}</div>
      <div class="field"><div class="flabel">Datasets used</div>${data}</div>
      ${note}${fileLink}
    </div>
  </div>`;
}
function haystack(p){return [p.std_name,p.title,p.summary,p.logical_flow,p.research_design,
  (p.authors||[]).join(" "),(p.categories||[]).join(" "),(p.datasets||[]).join(" "),p.journal].join(" ").toLowerCase();}

function renderPapers(){
  const q=document.getElementById("q").value.trim().toLowerCase();
  const fc=document.getElementById("fcat").value, fj=document.getElementById("fjour").value, fy=document.getElementById("fyear").value;
  const out=papers.filter(p=>(!q||haystack(p).includes(q))
    &&(!fc||p.categories.includes(fc))&&(!fj||p.journal===fj)&&(!fy||String(p.year)===fy));
  document.getElementById("papers").innerHTML = out.length?out.map(paperCard).join(""):`<div class="empty">No papers match. Try a broader search.</div>`;
  document.getElementById("pcount").textContent = out.length+" / "+papers.length+" papers";
}
function dsCard(d){
  const name = esc(d.provider)+(d.product?` — ${esc(d.product)}`:"");
  const tags = (d.topic_tags||"").split(";").map(t=>t.trim()).filter(Boolean).map(t=>`<span class="chip data">${esc(t)}</span>`).join("");
  const bits = [d.access_type,d.delivery].filter(Boolean).map(esc).join(" · ");
  const compiled = d.sources>1? `<div class="meta">description compiled from ${d.sources} papers</div>`:"";
  const used = d.used_by.length? `<div class="used"><b>${d.used_by.length}</b> paper(s): ${d.used_by.map(esc).join(" · ")}</div>`:"";
  const alias = d.aliases&&d.aliases!=="—"? `<div class="meta">aka ${esc(d.aliases)}</div>`:"";
  return `<div class="card">
    <div class="ptitle" style="cursor:default">${name}</div>
    ${alias}<div class="field">${esc(d.description||"")}</div>
    ${compiled}${bits?`<div class="meta">${bits}</div>`:""}${tags?`<div class="cats">${tags}</div>`:""}${used}
  </div>`;
}
// ---- broad theme -> narrow tag filter for datasets ----
// ordered most-specific -> most-general: the first matching theme wins
const THEME_DEFS = [
  ["Health Care", ["medicare","cms","dmepos","hhs","health care","hospital"]],
  ["Patents & Innovation", ["patent","uspto","kpss","post-grant","innovation","r&d"]],
  ["Payments", ["interchange","merchant acquir","fiserv","nilson","payment","dcpc"]],
  ["Private Markets (PE/VC)", ["preqin","pitchbook","crunchbase","venture","private equity","private credit","cap table","cem benchmarking","failed fund","capital commitment","limited partner","general partner","irr"]],
  ["Municipal, Housing & Regional", ["municipal","msrb","bond spread","house price","fhfa","lodes","property value","regional"]],
  ["Labor & Employment", ["employer-employee","occupation","employment","workforce","job posting","isco","wage","monopsony","lehd","executive survey","labor"]],
  ["Prices, Trading & Sentiment", ["stock","trading","market cap","sentiment","s-score","return","insider","congressional","social media","informativeness"]],
  ["Institutional Holdings", ["13f","institutional investor","intermediary","fire sale","non-bank","factset","morningstar"]],
  ["Banking & Deposits", ["call report","fdic","deposits","overdraft","nsf","branch","bank run","commercial bank"]],
  ["Credit & Lending", ["dealscan","covenant","syndicated","debtor-in-possession","dip","y-14","loan","interest rate","borrow","lending","credit","default","recovery rate"]],
  ["Legal & Enforcement", ["doj","litigation","indictment","exclusion","fraud enforcement","fraud charge","whistleblower","court record"]],
  ["Surveys", ["survey","questionnaire","csrc","diary","respondent","gallup","livingston","confidence"]],
  ["Corporate Filings & Text", ["10-k","edgar","sec analytics","newspaper","chronicling","text-based"]],
  ["Macro & Regulatory", ["manufacturing","nber-ces","sloos","census","macroeconomic","regulation","upia","fiduciary","prudent","lending standards"]],
];
function dsHay(d){return [d.provider,d.product,d.description,d.topic_tags,d.aliases].join(" ").toLowerCase();}
function themeOf(d){const h=dsHay(d);for(const [name,kws] of THEME_DEFS){if(kws.some(k=>h.includes(k)))return name;}return "Other";}
datasets.forEach(d=>{d._theme=themeOf(d); d._tags=(d.topic_tags||"").split(";").map(t=>t.trim()).filter(Boolean);});
let activeTheme=null, activeTag=null;
function renderThemeBar(){
  const counts={}; datasets.forEach(d=>counts[d._theme]=(counts[d._theme]||0)+1);
  const order=THEME_DEFS.map(t=>t[0]).concat(["Other"]).filter(n=>counts[n]);
  const bar=document.getElementById("dthemes");
  bar.innerHTML='<span class="barlabel">Theme</span>'+order.map(n=>
    `<span class="fchip${activeTheme===n?' on':''}" data-theme="${esc(n)}">${esc(n)}<span class="n">${counts[n]}</span></span>`).join("");
  bar.querySelectorAll(".fchip").forEach(c=>c.addEventListener("click",()=>{
    const n=c.dataset.theme; activeTheme=activeTheme===n?null:n; activeTag=null;
    renderThemeBar(); renderTagBar(); renderDatasets();
  }));
}
function renderTagBar(){
  const bar=document.getElementById("dtags");
  if(!activeTheme){bar.innerHTML="";return;}
  const tags={}; datasets.filter(d=>d._theme===activeTheme).forEach(d=>d._tags.forEach(t=>tags[t]=(tags[t]||0)+1));
  const order=Object.keys(tags).sort((a,b)=>tags[b]-tags[a]||a.localeCompare(b));
  if(!order.length){bar.innerHTML="";return;}
  bar.innerHTML='<span class="barlabel">Narrow</span>'+order.map(t=>
    `<span class="fchip${activeTag===t?' on':''}" data-tag="${esc(t)}">${esc(t)}<span class="n">${tags[t]}</span></span>`).join("");
  bar.querySelectorAll(".fchip").forEach(c=>c.addEventListener("click",()=>{
    const t=c.dataset.tag; activeTag=activeTag===t?null:t; renderTagBar(); renderDatasets();
  }));
}
function renderDatasets(){
  const q=document.getElementById("dq").value.trim().toLowerCase();
  const out=datasets.filter(d=>
    (!q||dsHay(d).includes(q)) &&
    (!activeTheme||d._theme===activeTheme) &&
    (!activeTag||d._tags.includes(activeTag)));
  document.getElementById("datasets").innerHTML = out.length?out.map(dsCard).join(""):`<div class="empty">No datasets match.</div>`;
  const active=[activeTheme,activeTag].filter(Boolean).length;
  document.getElementById("dcount").textContent = out.length+" / "+datasets.length+" datasets"+(active?" (filtered)":"");
}
["q","fcat","fjour","fyear"].forEach(id=>document.getElementById(id).addEventListener("input",renderPapers));
document.getElementById("dq").addEventListener("input",renderDatasets);
document.querySelectorAll(".tab").forEach(t=>t.addEventListener("click",()=>{
  document.querySelectorAll(".tab").forEach(x=>x.classList.remove("on")); t.classList.add("on");
  const isP=t.dataset.tab==="papers";
  document.getElementById("papers-view").style.display=isP?"":"none";
  document.getElementById("datasets-view").style.display=isP?"none":"";
}));
renderPapers(); renderThemeBar(); renderTagBar(); renderDatasets();
</script></body></html>"""

# --------------------------------------------------------------------------- #
# SEED  (loads your two sample summaries + four sample datasets as a live demo)
# --------------------------------------------------------------------------- #
def seed():
    conn = get_conn()
    for rec in _SAMPLE_PAPERS:
        rec = dict(rec)
        rec["std_name"] = build_std_name(rec)
        h = hashlib.sha256(("SEED::" + rec["title"]).encode()).hexdigest()
        if conn.execute("SELECT 1 FROM papers WHERE file_hash=?", (h,)).fetchone():
            continue
        save_paper(conn, rec, h, f"(sample — no file) {rec['title']}", "sample")
    for d in _SAMPLE_DATASETS:
        upsert_dataset(conn, d)
    conn.commit(); conn.close()
    print("Seeded sample papers and datasets.")
    build()

def refine():
    """Re-synthesize every dataset's description from its accumulated snippets.
    Uses the Claude API; run it occasionally rather than on every ingest."""
    client = _client()
    conn = get_conn()
    rows = conn.execute("SELECT * FROM datasets").fetchall()
    n = 0
    for d in rows:
        obs = json.loads(d["observations"] or "[]")
        if len(obs) < 1:
            continue
        name = d["provider"] + (f" — {d['product']}" if d["product"] else "")
        try:
            desc = synthesize_description(name, _dedup_snippets(obs), client) if len(obs) > 1 else obs[0]
            conn.execute("UPDATE datasets SET description=? WHERE id=?", (desc, d["id"]))
            n += 1
            print(f"  refined: {name[:60]}")
        except Exception as e:
            print(f"  skipped {name[:40]}: {str(e)[:60]}")
    conn.commit(); conn.close()
    print(f"Refined {n} descriptions.")
    build()

def stats():
    conn = get_conn()
    p = conn.execute("SELECT COUNT(*) c FROM papers").fetchone()["c"]
    d = conn.execute("SELECT COUNT(*) c FROM datasets").fetchone()["c"]
    l = conn.execute("SELECT COUNT(*) c FROM paper_datasets").fetchone()["c"]
    conn.close()
    print(f"{p} papers, {d} datasets, {l} paper-dataset links.\nDB: {DB_PATH}")

_SAMPLE_PAPERS = [
    {"journal":"American Economic Review","is_working_paper":False,"year":2003,
     "authors":["Antweiler"],
     "title":"How Effective is Green Regulatory Threat? Empirical Evidence from Canadian Plant-Level Data",
     "summary":"This paper investigates whether the mere threat of environmental regulation is effective in reducing industrial emissions. Using Canadian plant-level data from the National Pollutant Release Inventory (1993-1999), the author builds a theoretical model linking regulatory threat to firms' abatement decisions. The model predicts responses depend on environmental exposure and a firm's position on the 'abatement ladder.' Empirically, regulatory threat has a statistically significant but very small effect on emissions. The study concludes that regulatory threat alone is unlikely to be an effective policy tool in Canada.",
     "logical_flow":"The paper defines regulatory threat as a tool where firms cut emissions to avoid future regulation, builds a model predicting responses from exposure and abatement costs, tests the predictions on NPRI data controlling for firm characteristics, finds the overall impact minimal, and closes with the implication that direct regulation or incentives may work better.",
     "research_design":"Combines a theoretical model with empirical analysis of plant-level panel data, using fixed-effects regressions and difference-in-differences to isolate the effect of regulatory threat while accounting for geography, firm size, and pollutant toxicity.",
     "categories":["Environmental Economics","Industrial Organization","Public Policy"],
     "datasets":[{"provider":"National Pollutant Release Inventory (NPRI)","product":None,
        "description":"Canadian plant-level pollutant release data, 1993-1999.",
        "access_type":"Public","delivery":None,"topic_tags":"emissions; plant-level; Canada"}],
     "no_nonstandard_datasets":False,"missing_notes":None},
    {"journal":"Journal of Accounting Research","is_working_paper":False,"year":2023,
     "authors":["Chow","et al."],
     "title":"Reciprocity in Corporate Tax Compliance: Evidence from Ozone Pollution",
     "summary":"This paper examines reciprocity in corporate tax compliance, arguing firms treat taxes as payment for public goods such as environmental quality. Using ground-level ozone data, firms headquartered in higher-ozone areas engage in more aggressive tax planning and lower effective tax rates. When ozone falls due to regulation, compliance improves. Effects are stronger in politically conservative areas and among stakeholder-oriented managers. The findings suggest firms reciprocate government provision of public goods.",
     "logical_flow":"Building on contractarian theory of the state, the authors argue poor air quality signals government inefficiency and lowers willingness to pay taxes, document a negative ozone-compliance relationship across U.S. firms and counties, use EPA-driven regulatory changes as a difference-in-differences source of causal evidence, and explore moderators like managerial stakeholder orientation and local politics.",
     "research_design":"Panel regressions plus a difference-in-differences design exploiting EPA nonattainment reclassifications that reduce ozone, linking pollution to cash effective tax rates.",
     "categories":["Corporate Taxation","Environmental Economics","Political Economy"],
     "datasets":[
        {"provider":"EPA","product":"Ground-level ozone monitoring","description":"County-level ozone pollution and nonattainment designations.","access_type":"Public","delivery":None,"topic_tags":"air quality; ozone; county-level"},
        {"provider":"Google Trends","product":None,"description":"Public attention to pollution used in robustness tests.","access_type":"Public","delivery":"Web","topic_tags":"attention; search interest"}],
     "no_nonstandard_datasets":False,"missing_notes":None},
]

_SAMPLE_DATASETS = [
    {"provider":"LinkedIn","product":"Company Followers API","aliases":None,
     "description":"API providing follower counts and engagement data for company pages; used in marketing and hiring analytics.",
     "access_type":"Public","delivery":"API","topic_tags":"followers; engagement; brand presence"},
    {"provider":"GDELT Project","product":"Global Database of Events, Language and Tone","aliases":"GDELT",
     "description":"Open global news-derived databases: event data coded from news, a global knowledge graph, and tone; multilingual and near real time.",
     "access_type":"Public","delivery":"API","topic_tags":"event data; knowledge graph; tone; multilingual"},
    {"provider":"Socrata","product":"Open Data API (SODA)","aliases":"SODA; SoQL",
     "description":"Programmatic access layer for many open-data portals: discovery, querying, and downloads via SODA APIs.",
     "access_type":"Public","delivery":"API","topic_tags":"open data portals; querying; bulk retrieval"},
    {"provider":"ProPublica","product":"Data Store & APIs","aliases":None,
     "description":"Public datasets released for reporting plus APIs including Congress and Nonprofit Explorer endpoints.",
     "access_type":"Public","delivery":"API","topic_tags":"Congress API; Nonprofit Explorer; datasets archive"},
]

# --------------------------------------------------------------------------- #
def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "help"
    if   cmd == "ingest": ingest()
    elif cmd == "build":  build()
    elif cmd == "seed":   seed()
    elif cmd == "refine": refine()
    elif cmd == "stats":  stats()
    else:
        print(__doc__)

if __name__ == "__main__":
    main()
